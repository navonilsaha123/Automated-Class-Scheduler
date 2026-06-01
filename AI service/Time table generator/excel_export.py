from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from constants import DAYS, LUNCH_BREAK
from models import DepartmentPreview


HEADER_FILL = PatternFill("solid", fgColor="DCE6F2")
LAB_FILL = PatternFill("solid", fgColor="D9EAD3")
FREE_FILL = PatternFill("solid", fgColor="F3F3F3")
LUNCH_FILL = PatternFill("solid", fgColor="EAD1DC")
TITLE_FILL = PatternFill("solid", fgColor="FCE5CD")


def export_timetable_workbook(
    output_path: Path, institute_name: str, departments: list[DepartmentPreview]
) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for department in departments:
        worksheet = workbook.create_sheet(_safe_sheet_name(department.departmentName))
        _write_department_sheet(worksheet, institute_name, department)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _safe_sheet_name(name: str) -> str:
    invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
    cleaned = name
    for char in invalid_chars:
        cleaned = cleaned.replace(char, "-")
    return cleaned[:31] or "Department"


def _write_department_sheet(worksheet, institute_name: str, department: DepartmentPreview) -> None:
    worksheet["A1"] = "Institute"
    worksheet["B1"] = institute_name
    worksheet["A2"] = "Department"
    worksheet["B2"] = department.departmentName

    for cell in ("A1", "A2", "B1", "B2"):
        worksheet[cell].font = Font(bold=True)
        worksheet[cell].fill = TITLE_FILL

    worksheet.append([])
    header_row = ["Slot", "Time", *DAYS]
    worksheet.append(header_row)
    for cell in worksheet[4]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row = 5
    for slot in department.slots:
        worksheet.cell(row=current_row, column=1, value=slot.slotCode)
        worksheet.cell(row=current_row, column=2, value=slot.time)

        for day_offset, day in enumerate(DAYS, start=3):
            value = slot.entries[day]
            cell = worksheet.cell(row=current_row, column=day_offset, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if "Lab" in value or "LAB" in value:
                cell.fill = LAB_FILL
            elif "Free / Self Study" in value:
                cell.fill = FREE_FILL

        current_row += 1

        if slot.slotCode == "P4":
            worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            lunch_cell = worksheet.cell(
                row=current_row,
                column=1,
                value=f"{LUNCH_BREAK['label']} - {LUNCH_BREAK['time']}",
            )
            lunch_cell.fill = LUNCH_FILL
            lunch_cell.font = Font(italic=True, bold=True)
            lunch_cell.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1

    for column in ("A", "B", "C", "D", "E", "F", "G"):
        worksheet.column_dimensions[column].width = 20
